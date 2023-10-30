from django.db.models.query_utils import Q
from Ugovori.settings import STATIC_ROOT
from django.shortcuts import get_object_or_404, redirect, render, reverse
from django.db.models import Max, Sum, Count, Prefetch
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, View
from django.http import HttpResponse, JsonResponse
from django.views import generic
from .models import Preduzece, Ugovor, VrstaVozila, User, PoslovnaJedinica, Gorivo, Banka, BankovniIzvod, StavkeBankovnogIzvoda
from .forms import  (UserForm, UgovorModelForma, UgovorModelFormaIzmjene, 
                     CompanyForm, VrstaVozilaForm, VrstaGorivaForm, PoslovnaJedinicaForm, BankaForm,
                     IzvodSubformSetCreate, IzvodSubformSet, UserFormUpdate, UserPasswordUpdate)
       
from .utils import render_to_pdf, number_to_letters
from django.template.loader import get_template
from Ugovori.settings import STATICFILES_DIRS
from pathlib import Path
from PyPDF2 import PdfFileMerger
import io
from django.db.models import F, Func

from datetime import datetime, date
from django.contrib import messages
from django.utils.dateparse import parse_date

from openpyxl.styles import PatternFill
from openpyxl import Workbook


class LandingPageView(LoginRequiredMixin, TemplateView):
    
    template_name= 'landing.html'

def landing_page(request):
    return render(request, 'landing.html')

class ListaUgovora(LoginRequiredMixin, generic.ListView):
    template_name='kupoprodaja/lista_ugovora.html'
    context_object_name='ugovori'
    paginate_by=10
    def get_queryset(self):
        queryset= Ugovor.objects.all()
        user= self.request.user
        if self.request.user.is_superuser:
            queryset= Ugovor.objects.filter(storno=False).order_by('-broj_ugovora')
        elif user.bussines_unit_head:
            korisnik=User.objects.filter(poslovna_jedinica=user.poslovna_jedinica)
            queryset= Ugovor.objects.filter(korisnik_id__in=korisnik).filter(storno=False).order_by('-broj_ugovora')
        else:
            queryset=queryset.filter(korisnik_id=self.request.user).filter(storno=False).order_by('-broj_ugovora')
        return queryset

class ListaZahtjevaPregled(LoginRequiredMixin, generic.ListView):
    template_name='kupoprodaja/view_requested.html'
    context_object_name='ugovori'
    paginate_by=10
    def get_queryset(self):
        queryset= Ugovor.objects.all()
        user= self.request.user
        if self.request.user.is_superuser:
            
            queryset= Ugovor.objects.filter(storno=False).filter(ugovor_zakljucen=True).filter(zahtjev_pregled=True).order_by('-broj_ugovora')
        elif user.bussines_unit_head:
            korisnik=User.objects.filter(poslovna_jedinica=user.poslovna_jedinica)
            queryset= Ugovor.objects.filter(korisnik_id__in=korisnik).filter(storno=False).filter(ugovor_zakljucen=True).filter(zahtjev_pregled=True).order_by('-broj_ugovora')
        else:
            queryset=queryset.filter(korisnik_id=self.request.user).filter(storno=False).filter(ugovor_zakljucen=True).filter(zahtjev_pregled=True).order_by('-broj_ugovora')
        return queryset
def search_view(request): 
    buyer_name = request.GET.get("ime_kupac")
    payload = []
    if request.user.is_superuser:
        if buyer_name:
            all_contracts = Ugovor.objects.filter(
                Q(ime_kupac__icontains=buyer_name) |
                Q(ime_komitent__icontains=buyer_name) |
                Q(broj_ugovora__icontains=buyer_name)
            ).order_by('-broj_ugovora').annotate(
                buyer_name=F('ime_kupac'),
            )
            for one_contract in all_contracts:
                payload.append({"pk": one_contract.pk,"item": f"{one_contract.broj_ugovora} - {one_contract.ime_kupac} ({one_contract.ime_komitent})"})
    else:
        if buyer_name:
            all_contracts = Ugovor.objects.filter(korisnik_id=request.user).filter(
            Q(ime_kupac__icontains=buyer_name) |
            Q(ime_komitent__icontains=buyer_name) |
             Q(broj_ugovora__icontains=buyer_name)
            ).order_by('-broj_ugovora').annotate(
            buyer_name=F('ime_kupac'),
            )

            for one_contract in all_contracts:
                payload.append({"pk": one_contract.pk,"item": f"{one_contract.broj_ugovora} - {one_contract.ime_kupac} ({one_contract.ime_komitent})"})
            #payload.append(str(one_contract.broj_ugovora) + "-"+ one_contract.ime_kupac + " (" + one_contract.ime_komitent +")")
    return JsonResponse({"status": 200, "data": payload}, safe=False)

def kreiraj_ugovor(request):
    if request.method == 'POST':
        form = UgovorModelForma(request.POST)
        #print(form.errors)
        if form.is_valid():
            firma = Preduzece.objects.get(id=1)
            instance = form.save(commit=False)
            instance.korisnik = request.user
            instance.business_unit = request.user.poslovna_jedinica
            instance.cijena_neto = instance.cijena / (1 + float(firma.stopa_pdv))
            instance.cijena_neto = round(instance.cijena_neto, 2)
            instance.pdv = instance.cijena_neto * float(firma.stopa_pdv)
            instance.pdv = round(instance.pdv, 2)

            if not Ugovor.objects.exists():
                instance.broj_ugovora = 1
            else:
                
                instance.broj_ugovora = Ugovor.objects.aggregate(Max('broj_ugovora')).get('broj_ugovora__max') + 1
            
            instance.save()
            messages.success(request, 'Ugovor uspješno kreiran')
            return redirect(reverse("kupoprodaja:izmjena-ugovora", kwargs={'pk': instance.pk}))

    else:
        form = UgovorModelForma()

    context = {
        'form': form,
    }

    return render(request, 'kupoprodaja/kreiraj_ugovor.html', context)

def izmjeni_ugovor(request, pk):
    ugovor = get_object_or_404(Ugovor, pk=pk)
    firma = Preduzece.objects.get(id=1)

    if request.method == 'POST':
        form = UgovorModelFormaIzmjene(request.POST, instance=ugovor)
        #print(form.errors)
        if form.is_valid():
            form.instance.cijena_neto = form.instance.cijena / (1 + float(firma.stopa_pdv))
            form.instance.cijena_neto = round(form.instance.cijena_neto, 2)
            form.instance.pdv = form.instance.cijena_neto * float(firma.stopa_pdv)
            form.instance.pdv = round(form.instance.pdv, 2)
            
            messages.success(request, "Uspješno izmjenjeno!")
            form.save()
            
            return redirect('kupoprodaja:izmjena-ugovora', pk=ugovor.pk)
    else:
        form = UgovorModelFormaIzmjene(instance=ugovor)

    return render(request, 'kupoprodaja/izmjena_ugovora.html', {'form': form, 'ugovor': ugovor})

def zakljuci_ugovor(request, pk):
    ugovor = get_object_or_404(Ugovor, pk=pk)
    
    if request.method == 'POST':
       
        ugovor.ugovor_zakljucen = not ugovor.ugovor_zakljucen
        
        ugovor.save()
    messages.success(request, 'Ugovor je uspješno zaključen.')
    return redirect("kupoprodaja:izmjena-ugovora", pk=pk)

def view_request(request, pk):
    ugovor = get_object_or_404(Ugovor, pk=pk)
    
    if request.method == 'POST':
       
        ugovor.zahtjev_pregled = True
        
        ugovor.save()
    messages.success(request, 'Zahtjev za pregled poslan')
    return redirect("kupoprodaja:izmjena-ugovora", pk=pk)

def view_occured(request, pk):
    ugovor = get_object_or_404(Ugovor, pk=pk)
    
    if request.method == 'POST':
        ugovor.ugovor_pregledan = True
       
        ugovor.save()
        
        return redirect("kupoprodaja:izmjena-ugovora", pk=pk)

def storno_request(request, pk):
    ugovor = get_object_or_404(Ugovor, pk=pk)
    
    if request.method == 'POST':
       
        ugovor.storno_request = not ugovor.storno_request
        
        ugovor.save()
    messages.success(request, 'Zahtjev za storno uspješno poslat.')
    return redirect("kupoprodaja:izmjena-ugovora", pk=pk)

class BrisiUgovor(LoginRequiredMixin, generic.DeleteView):
    template_name='kupoprodaja/brisanje_ugovora.html'

    def get_queryset(self):
        queryset= Ugovor.objects.all()
        if self.request.user.is_superuser:
            queryset= Ugovor.objects.all()
        else:
            queryset=queryset.filter(korisnik_id=self.request.user)
        return queryset
    def get_success_url(self):
        return reverse('kupoprodaja:lista-ugovora')

class KreirajKorisnika(LoginRequiredMixin, generic.CreateView):
    model=User
    form_class=UserForm
    template_name = 'kupoprodaja/create_user.html'
 
    def form_valid(self, form):
        user = form.save(commit=False)
        user.set_password(user.password)
        user.save()
        return super(KreirajKorisnika, self).form_valid(form)
    def get_success_url(self):
        return reverse('kupoprodaja:lista-korisnika')

class ListaKorisnika(LoginRequiredMixin, generic.ListView):
    template_name='kupoprodaja/lista_korisnika.html'
    context_object_name='users'

    def get_queryset(self):
        queryset = User.objects.all()
        return queryset

class IzmjeniKorisnika(LoginRequiredMixin, generic.UpdateView):
    template_name='kupoprodaja/izmjena_korisnika.html'
    form_class = UserFormUpdate

    def get_queryset(self):
        queryset= User.objects.all()
        return queryset
    def form_valid(self, form):
        user = form.save(commit=False)
        user.save()
        return super(IzmjeniKorisnika, self).form_valid(form)
    def get_success_url(self):
        return reverse("kupoprodaja:lista-korisnika")

class IzmjeniSifruKorisnika(LoginRequiredMixin, generic.UpdateView):
    template_name='kupoprodaja/izmjena_sifre.html'
    form_class = UserPasswordUpdate

    
    def get_queryset(self):
        queryset= User.objects.all()
        return queryset
    def form_valid(self, form):
        user = form.save(commit=False)
        user.set_password(user.password)
        user.save()
        messages.success(self.request, 'Šifra uspješno promjenjena!')
        return super(IzmjeniSifruKorisnika, self).form_valid(form)
    def get_success_url(self):
        return reverse("kupoprodaja:lista-ugovora")
class GeneratePDFContract(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        ugovor = get_object_or_404(Ugovor, pk=pk)
        decimalni_broj = ((ugovor.vrijednost_vozila) - int(ugovor.vrijednost_vozila)) * 100
        vrijednost = number_to_letters(ugovor.vrijednost_vozila)
        if request.user.is_superuser == False:
            if ugovor.ugovor_pregledan==True:
                messages.error(request, 'Ugovor je već štampan, pošaljite zahtjev za štampu')
                return redirect("kupoprodaja:izmjena-ugovora", pk=pk)       
      
    
        preduzece = get_object_or_404(Preduzece)
        BASE_DIR = Path(__file__).resolve().parent.parent
        
        # Create in-memory binary streams of the PDF content with different sentences
        pdf1_context = {
            'base_dir': BASE_DIR,
            'ugovor': ugovor,
            'firma': preduzece,
            'vrijednost': vrijednost,
            'decimalni_broj': decimalni_broj,
            'order': 1,
            'unique_sentence': "ZA MUP PRIJAVA VOZILA",
        }
        pdf1_bytes = render_to_pdf('kupoprodaja/contract.html', pdf1_context).content

        pdf2_context = {
            'base_dir': BASE_DIR,
            'ugovor': ugovor,
            'firma': preduzece,
            'vrijednost': vrijednost,
            'decimalni_broj': decimalni_broj,
            'order': 2,
            'unique_sentence': "ZA KOMISIONARA",
        }
        pdf2_bytes = render_to_pdf('kupoprodaja/contract.html', pdf2_context).content

        pdf3_context = {
            'base_dir': BASE_DIR,
            'ugovor': ugovor,
            'firma': preduzece,
            'vrijednost': vrijednost,
            'decimalni_broj': decimalni_broj,
            'order': 3,
            'unique_sentence': "ZA KOMITENTA",
        }
        pdf3_bytes = render_to_pdf('kupoprodaja/contract.html', pdf3_context).content

        pdf4_context = {
            'base_dir': BASE_DIR,
            'ugovor': ugovor,
            'firma': preduzece,
            'vrijednost': vrijednost,
            'decimalni_broj': decimalni_broj,
            'order': 4,
            'unique_sentence': "ZA KUPCA",
        }
        pdf4_bytes = render_to_pdf('kupoprodaja/contract.html', pdf4_context).content
        
        # Merge PDFs
        merger = PdfFileMerger()
        merger.append(io.BytesIO(pdf1_bytes))
        merger.append(io.BytesIO(pdf2_bytes))
        merger.append(io.BytesIO(pdf3_bytes))
        merger.append(io.BytesIO(pdf4_bytes))
        
        # Output merged PDF to response
        response = HttpResponse(content_type='application/pdf')
        filename = "Ugovor_%s.pdf" % ("12341231")
        content_disposition = "inline; filename='%s'" % (filename)
        download = request.GET.get("download")
        if download:
            content_disposition = "attachment; filename='%s'" % (filename)
        
        # Write merged PDF to response
        merger.write(response)
        merger.close()  # Close the merger object
        
        response['Content-Disposition'] = content_disposition
        return response


class ContractStatistics(LoginRequiredMixin, generic.ListView):
    template_name='kupoprodaja/storno_contract_status.html'
    def get_context_data(self, **kwargs):
        context = super(ContractStatistics, self).get_context_data(**kwargs)
        queryset= Ugovor.objects.all()
        user= self.request.user
        if self.request.user.is_superuser:
            queryset= Ugovor.objects.all()
        elif user.bussines_unit_head:
            korisnik=User.objects.filter(poslovna_jedinica=user.poslovna_jedinica)
            queryset= Ugovor.objects.filter(korisnik_id__in=korisnik)
        else:
            queryset=queryset.filter(korisnik_id=self.request.user)
        context.update({
            "contract_count": queryset.filter(storno=False).filter(ugovor_zakljucen=True).count(),
            "contract_storno": queryset.filter(storno=True).count(),
            "contract_draft": queryset.filter(storno=False).filter(ugovor_zakljucen=False).count(),
            "storno_waiting": queryset.filter(storno=False).filter(storno_request=True).count(),
            "view_request": queryset.filter(storno=False).filter(ugovor_zakljucen=True).filter(zahtjev_pregled=True).count()
        }) 
        return context
    
    def get_queryset(self):
        queryset= Ugovor.objects.all()
        user= self.request.user
        if self.request.user.is_superuser:
            queryset= Ugovor.objects.all().order_by('-datum')
        elif user.bussines_unit_head:
            korisnik=User.objects.filter(poslovna_jedinica=user.poslovna_jedinica)
            queryset= Ugovor.objects.filter(korisnik_id__in=korisnik).order_by('-datum')
        else:
            queryset=queryset.filter(korisnik_id=self.request.user).order_by('-datum')
        return queryset

class ContractStornoRequested(LoginRequiredMixin, generic.ListView):
    template_name='kupoprodaja/storno_contract_requested.html' 
    context_object_name ="ugovori"
    paginate_by=10
    def get_queryset(self):
        queryset= Ugovor.objects.all()
        user= self.request.user
        if self.request.user.is_superuser:
            queryset= Ugovor.objects.filter(storno=False).filter(storno_request=True).order_by('-datum')
            
        elif user.bussines_unit_head:
            korisnik=User.objects.filter(poslovna_jedinica=user.poslovna_jedinica)
            queryset= Ugovor.objects.filter(korisnik_id__in=korisnik).filter(storno=False).filter(storno_request=True).order_by('-datum')
           
        else:
            queryset=queryset.filter(korisnik_id=self.request.user).filter(storno=False).filter(storno_request=True).order_by('-datum')
            
        return queryset

class ContractStornoApproved(LoginRequiredMixin, generic.ListView):
    template_name='kupoprodaja/storno_contract_approved.html' 
    context_object_name ="ugovori"
    paginate_by=10
    def get_queryset(self):
        queryset= Ugovor.objects.all()
        user= self.request.user
        if self.request.user.is_superuser:
            queryset= Ugovor.objects.filter(storno=True).order_by('-datum')
        elif user.bussines_unit_head:
            korisnik=User.objects.filter(poslovna_jedinica=user.poslovna_jedinica)
            queryset= Ugovor.objects.filter(korisnik_id__in=korisnik).filter(storno=True).order_by('-datum')
        else:
            queryset=queryset.filter(korisnik_id=self.request.user).filter(storno=True).order_by('-datum')
        
        return queryset

class ContractDraft(LoginRequiredMixin, generic.ListView):
    template_name='kupoprodaja/contract_draft.html' 
    context_object_name ="ugovori"
    paginate_by=10
    def get_queryset(self):
        queryset= Ugovor.objects.all()
        user= self.request.user
        if self.request.user.is_superuser:
            queryset= Ugovor.objects.filter(storno=False).filter(ugovor_zakljucen=False).order_by('-datum')
        elif user.bussines_unit_head:
            korisnik=User.objects.filter(poslovna_jedinica=user.poslovna_jedinica)
            queryset= Ugovor.objects.filter(korisnik_id__in=korisnik).filter(storno=False).filter(ugovor_zakljucen=False).order_by('-datum')
        else:
            queryset=queryset.filter(korisnik_id=self.request.user).filter(storno=False).filter(ugovor_zakljucen=False).order_by('-datum')
        
        return queryset
    

def contract_storno_approve(request, pk):
    ugovor = get_object_or_404(Ugovor, pk=pk)
    
    if request.method == 'POST':
       
        ugovor.storno = True
        
        ugovor.save()
    messages.success(request, 'Ugovor uspješno storniran.')
    return redirect("kupoprodaja:storno-zahtjevi")

def contract_storno_discard(request, pk):
    ugovor = get_object_or_404(Ugovor, pk=pk)
    
    if request.method == 'POST':
       
        ugovor.storno = False
        ugovor.storno_request = False
        ugovor.save()
    messages.error(request, 'Odbijen storno ugovora.')
    return redirect("kupoprodaja:storno-zahtjevi")

def contract_view_approve(request, pk):
    ugovor = get_object_or_404(Ugovor, pk=pk)
    
    if request.method == 'POST':
       
        ugovor.ugovor_pregledan = False
        ugovor.zahtjev_pregled = False
        
        ugovor.save()
    messages.success(request, 'Dozvoljen ponovni pregled ugovora.')
    return redirect("kupoprodaja:lista-pregled")

def contract_view_discard(request, pk):
    ugovor = get_object_or_404(Ugovor, pk=pk)
    
    if request.method == 'POST':
       
        ugovor.zahtjev_pregled = False
        ugovor.save()
    messages.error(request, 'Odbijen ponovni pregled ugovora.')
    return redirect("kupoprodaja:lista-pregled")

class Round(Func):
      function = 'ROUND'
      arity = 2

def is_valid_param(param):
    return param !='' and param is not None
         
class GenerateInvoiceList(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        BASE_DIR = Path(__file__).resolve().parent.parent
        user= self.request.user
        korisnik=User.objects.filter(poslovna_jedinica=user.poslovna_jedinica).values('poslovna_jedinica')
        ugovor=request.session['ugovor']
        if user.is_superuser:
            ugovor = ugovor
        else:
            ugovor=ugovor.filter(korisnik=user.id)
        startdate=parse_date(request.session['startdate'])
        enddate=parse_date(request.session['enddate'])
        totalcijena = 0.00
        totalpdv=0.00
        totalukupno=0.00
        if user.is_superuser:
            poslovnice= PoslovnaJedinica.objects.annotate(pdv=Sum('ugovor__pdv', filter=Q(ugovor__in=ugovor)),\
            cijena=Sum('ugovor__cijena', filter=Q(ugovor__in=ugovor)), broj=Count('ugovor__cijena', filter=Q(ugovor__in=ugovor)),\
            cijena_neto=Sum('ugovor__cijena_neto', filter=Q(ugovor__in=ugovor))).prefetch_related(Prefetch('ugovor_set',
            ugovor, to_attr='stavke'))
        else:
            poslovnice= PoslovnaJedinica.objects.filter(id__in=korisnik).annotate(pdv=Sum('ugovor__pdv', filter=Q(ugovor__in=ugovor)),\
            cijena=Sum('ugovor__cijena', filter=Q(ugovor__in=ugovor)), broj=Count('ugovor__cijena', filter=Q(ugovor__in=ugovor)),\
            cijena_neto=Sum('ugovor__cijena_neto', filter=Q(ugovor__in=ugovor))).prefetch_related(Prefetch('ugovor_set',
            ugovor, to_attr='stavke'))

        
        totalcijena = ugovor.aggregate(Sum('cijena_neto'))
        totalcijena = {k: round(v, 2) for k, v in totalcijena.items()}
        totalbroj = ugovor.aggregate(Count('pdv'))
        totalpdv = ugovor.aggregate(Sum('pdv'))
        totalpdv = {k: round(v, 2) for k, v in totalpdv.items()}
        totalukupno = ugovor.aggregate(Sum('cijena'))
        totalukupno = {k: round(v, 2) for k, v in totalukupno.items()}

        template = get_template('kupoprodaja/invoice_list.html')
        preduzece = get_object_or_404(Preduzece)
        context = {
                'base_dir': BASE_DIR,
                'poslovnice': poslovnice,
                'start': startdate,
                'end': enddate,
                'firma': preduzece,             
                'totalukupno':totalukupno,
                'totalcijena': totalcijena,
                'totalpdv': totalpdv,
                'totalbroj': totalbroj,
                'user': user
            }
        html = template.render(context)
        pdf = render_to_pdf('kupoprodaja/invoice_list.html', context)
        
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "KIF_%s.pdf" %("12341231")
            content = "inline; filename='%s'" %(filename)
            download = request.GET.get("download")
            if download:
                content = "attachment; filename='%s'" %(filename)
            response['Content-Disposition'] = content
            return response
        return HttpResponse("Not found")
    
def ReportView(request):
    pregled=request.GET.get('Report_type')
    user=request.user
    ugovor= Ugovor.objects.filter(storno=False).filter(ugovor_zakljucen=True)
    korisnik = User.objects.all()
    usr= User.objects.all()
    if request.user.is_superuser:
        ugovor= ugovor
    else:
        korisnik=User.objects.filter(poslovna_jedinica=user.poslovna_jedinica)
        ugovor= ugovor.filter(korisnik_id__in=korisnik)
    startdate=request.GET.get('start_date')
    enddate=request.GET.get('end_date')
    user_filter = request.GET.get('User_selected')
    if is_valid_param(startdate):
        ugovor=ugovor.filter(datum__gte=(startdate))
    if is_valid_param(enddate):
        ugovor=ugovor.filter(datum__lte=(enddate))
    request.session['ugovor']=ugovor
    request.session['startdate']=startdate
    request.session['enddate']=enddate
    request.session['korisnik']=user_filter
   
    if pregled == 'kif':
        return GenerateInvoiceList.as_view()(request)
    elif pregled =='uplate':
        return redirect('kupoprodaja:eksport')
    elif pregled =='fakture':
        return redirect('invoice:rekapitulacija')
    elif pregled =='uplate_fakture':
        return redirect('invoice:eksport-fakture')
    elif pregled =='rekapitulacija_zbirno':
        return redirect('kupoprodaja:rekapitulacija-zbirno')
    elif pregled =='rekapitulacija_pojedinacno':
        return redirect('kupoprodaja:rekapitulacija-pojedinacno')
    return render(request, 'kupoprodaja/reports.html', {'ugovori': ugovor, 'usr': usr}) 
def contract_payment(request, pk):

    ugovor = get_object_or_404(Ugovor, pk=pk)
    
    if request.method == 'POST':
       
        ugovor.naplaceno = 'DA'
        
        ugovor.save()

    messages.success(request, 'Naplata ugovora potvrđena')
    return redirect("kupoprodaja:izmjena-ugovora", pk=pk)
def contract_payment_list(request, pk):
    ugovor = get_object_or_404(Ugovor, pk=pk)
    current_page = request.GET.get('pageTemp')

    if request.method == 'POST':
        # Process your form data and update the messages
        ugovor.naplaceno = 'DA'
        ugovor.save()

        messages.success(request, 'Naplata ugovora potvrđena')
        messages_data = [{'message': str(message), 'class': 'json-message'} for message in messages.get_messages(request)]

        if current_page is not None:
            redirect_url = reverse('kupoprodaja:lista-ugovora') + f'?page={current_page}'
        else:
            redirect_url = reverse('kupoprodaja:lista-ugovora')

        if request.is_ajax():
            
            return JsonResponse({'status': 'success', 'messages': messages_data})

        return redirect(redirect_url)

    return redirect('kupoprodaja:lista-ugovora')
   
def export_excel(request):
    user = request.user
    ugovor=request.session['ugovor']
    if user.is_superuser:
        ugovor = ugovor
    else:
        ugovor=ugovor.filter(korisnik=user.id)
    ugovor=ugovor.values('broj_ugovora', 'ime_kupac','jib_kupac', 'datum', 'cijena', 'naplaceno',  'korisnik__username')
   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-ugovori.xlsx'.format(
        date=datetime.now().strftime('%d.%m.%Y'),
    )
    workbook = Workbook()
    
    # Pozicija aktivnog sheet-a
    worksheet = workbook.active
    worksheet.title = 'Movies'

    # Definisanje naziva kolona
    columns = [
        'Ugovor', 'Kupac', 'JIB/JMBG', 'Datum','Cijena','Naplaćeno', 'Korisnik'
    ]
    row_num = 1

    # Upis naziva kolona u fajl
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iteracija kroz ugovore
    for x in ugovor:
        row_num += 1
        
        # definisanje podataka za pojedinačni red 
        row = [
            x['broj_ugovora'],
            x['ime_kupac'],
            x['jib_kupac'],
            x['datum'],
            x['cijena'],
            x['naplaceno'],
            #x['broker_id__ime'],
            x['korisnik__username'],
        ]
        
        # upis podataka u redove
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
        
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 10
    worksheet.column_dimensions['F'].width = 10
    #worksheet.column_dimensions['G'].width = 18
    worksheet.column_dimensions['H'].width = 18
    
    # formatiranje ćelija
    rows=worksheet.max_row
    
    for i in range(2, rows+1):
        cell= 'D' + f'{i}'
        date_item=worksheet[cell].value
        date_item=date_item.strftime('%d.%m.%Y')
        worksheet[cell].value=date_item
    
    fill_red = PatternFill(start_color='FF5D3A', end_color='FF5D3A', fill_type='solid')
    fill_green = PatternFill(start_color='29CE08', end_color='29CE08', fill_type='solid')

    for i in range(2, rows+1):
        cell= 'F' + f'{i}'
        paid_item = worksheet[cell]
        if paid_item.value == 'DA':
            paid_item.fill = fill_green
        else:
            paid_item.fill = fill_red
    workbook.save(response)
    
    return response



def bussines_unit_list(request):
    
    bussines_units=PoslovnaJedinica.objects.all()
    
    return render(request, "kupoprodaja/lista_poslovnica.html", {"bussines_units": bussines_units})



def vehicle_type_list(request):
    
    vehicle_types=VrstaVozila.objects.all()
   
    return render(request, "kupoprodaja/vrste_vozila.html", {"vehicle_types": vehicle_types})

def fuel_type_list(request):
    
    fuel_types=Gorivo.objects.all()
   
    return render(request, "kupoprodaja/vrste_goriva.html", {"fuel_types": fuel_types})

def company_details(request):
    company_form = get_object_or_404(Preduzece)
    if request.method== "POST":
        
        company_form=CompanyForm(request.POST, instance=company_form)

        if company_form.is_valid():
            post=company_form.save(commit=False)
            post.save()
            messages.success(request, 'Uspješno izmjenjeno')
            return redirect("kupoprodaja:podaci-firma")
        #else:
            #return render(request, "kupoprodaja/firma.html", {"company_form": company_form})
    
     
    elif request.method=="GET":
        company_form=CompanyForm(instance=company_form)
        return render(request, 'kupoprodaja/firma.html', {"company_form": company_form})
    
    return render(request, "kupoprodaja/firma.html", {"company_form": company_form})


def kreiraj_vrstu_vozila(request):
    if request.method == 'POST':
        form = VrstaVozilaForm(request.POST)
        #print(form.errors)
        if form.is_valid():       
            instance = form.save(commit=False)
            instance.vrsta_vozila = instance.vrsta_vozila           
            instance.save()
            messages.success(request, 'Vozilo uspješno kreirano')
            return redirect(reverse("kupoprodaja:izmjena-vrste-vozila", kwargs={'pk': instance.pk}))
    else:
        form = VrstaVozilaForm()

    context = {
        'form': form,
    }
    return render(request, 'kupoprodaja/vrsta_vozila_kreiraj.html', context)

def izmjeni_vrstu_vozila(request, pk):
    vrsta = get_object_or_404(VrstaVozila, pk=pk)
    if request.method == 'POST':
        form = VrstaVozilaForm(request.POST, instance=vrsta)
        #print(form.errors)
        if form.is_valid():
            form.instance.vrsta_vozila = form.instance.vrsta_vozila
           
            messages.success(request, "Uspješno izmjenjeno!")
            form.save()
            
            return redirect('kupoprodaja:izmjena-vrste-vozila', pk=vrsta.pk)
    else:
        form = VrstaVozilaForm(instance=vrsta)

    return render(request, 'kupoprodaja/vrsta_vozila_izmjena.html', {'form': form, 'vrsta': vrsta})

def kreiraj_vrstu_goriva(request):
    if request.method == 'POST':
        form = VrstaGorivaForm(request.POST)
        #print(form.errors)
        if form.is_valid():       
            instance = form.save(commit=False)
            instance.tip_goriva = instance.tip_goriva        
            instance.save()
            messages.success(request, 'Vrsta goriva uspješno kreirana')
            return redirect(reverse("kupoprodaja:izmjena-vrste-goriva", kwargs={'pk': instance.pk}))
    else:
        form = VrstaGorivaForm()

    context = {
        'form': form,
    }
    return render(request, 'kupoprodaja/vrsta_goriva_kreiraj.html', context)

def izmjeni_vrstu_goriva(request, pk):
    vrsta = get_object_or_404(Gorivo, pk=pk)
    if request.method == 'POST':
        form = VrstaGorivaForm(request.POST, instance=vrsta)
        if form.is_valid():
            form.instance.tip_goriva = form.instance.tip_goriva
           
            messages.success(request, "Uspješno izmjenjeno!")
            form.save()
            
            return redirect('kupoprodaja:izmjena-vrste-goriva', pk=vrsta.pk)
    else:
        form = VrstaGorivaForm(instance=vrsta)

    return render(request, 'kupoprodaja/vrsta_goriva_izmjena.html', {'form': form, 'vrsta': vrsta})


def kreiraj_poslovnicu(request):
    if request.method == 'POST':
        form = PoslovnaJedinicaForm(request.POST)
        
        if form.is_valid():       
            instance = form.save(commit=False)
            instance.naziv = instance.naziv       
            instance.save()
            messages.success(request, 'Poslovnica uspješno kreirana')
            return redirect(reverse("kupoprodaja:izmjena-poslovnice", kwargs={'pk': instance.pk}))
    else:
        form = PoslovnaJedinicaForm()

    context = {
        'form': form,
    }
    return render(request, 'kupoprodaja/poslovnica_kreiraj.html', context)

def izmjeni_poslovnicu(request, pk):
    poslovnica= get_object_or_404(PoslovnaJedinica, pk=pk)
    if request.method == 'POST':
        form = PoslovnaJedinicaForm(request.POST, instance=poslovnica)
        if form.is_valid():
            form.instance.naziv = form.instance.naziv
           
            messages.success(request, "Uspješno izmjenjeno!")
            form.save()
            
            return redirect('kupoprodaja:izmjena-poslovnice', pk=poslovnica.pk)
    else:
        form = PoslovnaJedinicaForm(instance=poslovnica)

    return render(request, 'kupoprodaja/poslovnica_izmjena.html', {'form': form, 'poslovnica': poslovnica})


def kreiraj_banku(request):
    if request.method == 'POST':
        form =  BankaForm(request.POST)
        #print(form.errors)
        if form.is_valid():       
            instance = form.save(commit=False)
                 
            instance.save()
            messages.success(request, 'Banka uspješno dodana')
            return redirect(reverse("kupoprodaja:izmjena-banka", kwargs={'pk': instance.pk}))
    else:
        form =  BankaForm()

    context = {
        'form': form,
    }
    return render(request, 'kupoprodaja/banka_kreiraj.html', context)

def izmjeni_banku(request, pk):
    banka = get_object_or_404(Banka, pk=pk)
    if request.method == 'POST':
        form =  BankaForm(request.POST, instance=banka)
        if form.is_valid():
            form.instance.naziv_banke = form.instance.naziv_banke
            form.instance.sjediste_banke = form.instance.sjediste_banke
            form.instance.adresa_banke = form.instance.adresa_banke
            messages.success(request, "Uspješno izmjenjeno!")
            form.save()
            
            return redirect('kupoprodaja:izmjena-banka', pk=banka.pk)
    else:
        form =  BankaForm(instance=banka)

    return render(request, 'kupoprodaja/banka_izmjena.html', {'form': form, 'banka': banka})


def banke_lista(request):
    
    lista_banaka=Banka.objects.all()
   
    return render(request, "kupoprodaja/banka_lista.html", {"lista_banaka": lista_banaka})

class IzvodKreirajFunkcija(LoginRequiredMixin, generic.CreateView):
    
    model= BankovniIzvod
    template_name='kupoprodaja/izvod_kreiranje.html'
    fields = [
            'broj_izvoda',
            'banka',
            'datum_izvoda',
            'napomena',
            'zakljucen_izvod',
        ]
   
    
    def get_context_data(self, **kwargs):
        # we need to overwrite get_context_data
        # to make sure that our formset is rendered
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data["items"] = IzvodSubformSetCreate(self.request.POST)
        else:
            data["items"] = IzvodSubformSetCreate()
        return data
    def form_valid(self, form):
        
        form.instance.poslovna_jedinica= self.request.user.poslovna_jedinica
        self.object = form.save()
        messages.success(self.request, 'Uspješno dodano')
        return super().form_valid(form)
    def get_success_url(self):
        return reverse("kupoprodaja:izvod-izmjena", kwargs={'pk': self.object.pk})
    
class IzvodIzmjeniFunkcija(LoginRequiredMixin, generic.UpdateView):
     model= BankovniIzvod
     template_name='kupoprodaja/izvod_izmjena.html'
     fields = [

            'broj_izvoda',
            'banka',
            'datum_izvoda',
            'napomena',
            'zakljucen_izvod',
        ]
   

     def get_context_data(self, **kwargs):

        data = super().get_context_data(**kwargs)
        if self.request.POST:
            
            data["items"] = IzvodSubformSet(self.request.POST, instance=self.object)
            
        else:
            data["items"] = IzvodSubformSet(instance=self.object)
        return data

     def form_valid(self, form):
        
        context = self.get_context_data()
        items = context["items"]
        
        if items.is_valid():
            
            for item in items:
                obj = item.save(commit=False)
                #usr = item.cleaned_data['korisnik_uplatio'].value()
                obj.korisnik_uplatio= item.cleaned_data['korisnik_uplatio']
                obj.opis = item.cleaned_data['opis']
                obj.uplata = item.cleaned_data['uplata']
                obj.isplata = item.cleaned_data['isplata'] 
                obj.save()
            messages.success(self.request, 'Uspješno sačuvano')
            
        else:
            errors=items.errors
            print(errors)
        return super().form_valid(form)

     def get_success_url(self):
            return reverse("kupoprodaja:izvod-izmjena", kwargs={'pk': self.object.pk})
        
def izvodi_lista(request):
    
    lista_izvoda=BankovniIzvod.objects.all().order_by("-broj_izvoda")
   
    return render(request, "kupoprodaja/izvodi_lista.html", {"lista_izvoda": lista_izvoda})



def export_dnevnik_pojedinacno(request):
    startdate = request.session['startdate']
    enddate = request.session['enddate']
    #request.session['korisnik']=korisnik
    if request.user.is_superuser:  
        user_id = request.session['korisnik']
    else:
        user_id=request.user.id
    BASE_DIR = Path(__file__).resolve().parent.parent
    preduzece = get_object_or_404(Preduzece)
    user_data = User.objects.filter(id=user_id)
    
    contracts_sum_valid = {'total_iznos': 0}  # Initialize with default value
    contracts_sum_storno = {'total_iznos_storno': 0}  # Initialize with default value
    contracts_sum_valid_prije = {'total_iznos_prije': 0}
    contracts_sum_storno_prije =  {'total_iznos_storno_prije': 0}
    uplata_sum = 0  # Initialize with default value
    isplata_sum = 0
    uplata_sum_prije=0

    results = None
    results_isplata = None

    if is_valid_param(startdate) or is_valid_param(enddate):
        contracts_sum_valid = Ugovor.objects.filter(
            storno=False,
            ugovor_zakljucen=True,
            korisnik=user_id,
            datum__gte=startdate,
            datum__lte=enddate
        ).aggregate(total_iznos=Sum('cijena'))
        
        contracts_sum_valid_prije = Ugovor.objects.filter(
            storno=False,
            ugovor_zakljucen=True,
            korisnik=user_id,
            datum__lt=startdate,

        ).aggregate(total_iznos_prije=Sum('cijena'))
        
        contracts_sum_storno = Ugovor.objects.filter(
            storno=True,
            ugovor_zakljucen=True,
            korisnik=user_id,
            datum__gte=startdate,
            datum__lte=enddate
        ).aggregate(total_iznos_storno=Sum('cijena'))
        
        contracts_sum_storno_prije = Ugovor.objects.filter(
            storno=True,
            ugovor_zakljucen=True,
            korisnik=user_id,
            datum__lt=startdate,

        ).aggregate(total_iznos_storno_prije=Sum('cijena'))

        uplata_sum = StavkeBankovnogIzvoda.objects.filter(
            id_bankovni_izvod__datum_izvoda__gte=startdate,
            id_bankovni_izvod__datum_izvoda__lte=enddate,
            korisnik_uplatio=user_id
        ).aggregate(total_uplata=Sum('uplata'))['total_uplata'] or 0
        
        uplata_sum_prije = StavkeBankovnogIzvoda.objects.filter(
          
            id_bankovni_izvod__datum_izvoda__lt=startdate,
            korisnik_uplatio=user_id
        ).aggregate(total_uplata=Sum('uplata'))['total_uplata'] or 0

        results = StavkeBankovnogIzvoda.objects.filter(
            id_bankovni_izvod__datum_izvoda__gte=startdate,
            id_bankovni_izvod__datum_izvoda__lte=enddate,
            korisnik_uplatio=user_id
        ).values('id_bankovni_izvod__banka__naziv_banke',
                 'id_bankovni_izvod__broj_izvoda',  # Renaming 'id_bankovni_izvod__broj_izvoda' field to 'broj_izvoda'
                 'id_bankovni_izvod__datum_izvoda',
                 'uplata',
                 'opis')
        
        isplata_sum = StavkeBankovnogIzvoda.objects.filter(
            id_bankovni_izvod__datum_izvoda__gte=startdate,
            id_bankovni_izvod__datum_izvoda__lte=enddate,
            korisnik_uplatio=user_id
        ).aggregate(total_isplata=Sum('isplata'))['total_isplata'] or 0
        


        results_isplata = StavkeBankovnogIzvoda.objects.filter(
            id_bankovni_izvod__datum_izvoda__gte=startdate,
            id_bankovni_izvod__datum_izvoda__lte=enddate,
            korisnik_uplatio=user_id
        ).values('id_bankovni_izvod__banka__naziv_banke',
                 'id_bankovni_izvod__broj_izvoda',  # Renaming 'id_bankovni_izvod__broj_izvoda' field to 'broj_izvoda'
                 'id_bankovni_izvod__datum_izvoda',
                 'isplata',
                 'opis')
    startdate=parse_date(startdate)
    enddate=parse_date(enddate)
    total_iznos_sum = contracts_sum_valid.get('total_iznos', 0)
    total_storno_sum = contracts_sum_storno.get('total_iznos_storno', 0)
    total_iznos_sum_prije = contracts_sum_valid_prije.get('total_iznos_prije', 0)
    total_storno_sum_prije = contracts_sum_storno_prije.get('total_iznos_storno_prije', 0)

    if total_iznos_sum is None:
        total_iznos_sum=0
    if total_storno_sum is None:
        total_storno_sum=0
    if total_iznos_sum_prije is None:
        total_iznos_sum_prije = 0
    if total_storno_sum_prije is None:
        total_storno_sum_prije = 0
    
    uplatiti=total_iznos_sum + total_iznos_sum_prije -float(uplata_sum) - float(uplata_sum_prije)
    

    context= {
        "user_data": user_data,
        "firma": preduzece,
        "total_iznos_sum": total_iznos_sum,
        "total_storno_sum": total_storno_sum,
        "uplata_sum": uplata_sum,
        "results": results,
        "isplata_sum": isplata_sum,
        "results_isplata": results_isplata,
        "base_dir": BASE_DIR,
        "uplatiti": uplatiti,
        "start": startdate,
        "end": enddate,
        "uplata_sum_prije": uplata_sum_prije,
        "total_iznos_sum_prije": total_iznos_sum_prije,

        "contracts_sum_storno_prije": total_storno_sum_prije
        
    }
    

    pdf = render_to_pdf('kupoprodaja/rekapitulacija_korisnik.html', context)
        
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = "rekapitulacija_%s.pdf" %("12341231")
        content = "inline; filename='%s'" %(filename)
        download = request.GET.get("download")
        if download:
            content = "attachment; filename='%s'" %(filename)
        response['Content-Disposition'] = content
        return response
    return HttpResponse("Not found")
    

def export_dnevnik_zbirno(request):
    startdate = request.session.get('startdate')
    enddate = request.session.get('enddate')
    user_ids = User.objects.all()
    #user_ids = user_ids.values_list('id', 'first_name', 'jmb_korisnika', 'broj_telefona', 'mjesto', 'adresa', 'licna_karta')


    BASE_DIR = Path(__file__).resolve().parent.parent
    preduzece = get_object_or_404(Preduzece)

    total_iznos_sum = 0
    total_storno_sum = 0
    uplata_sum = 0
    isplata_sum = 0
    isplata_sum_prije = 0
    total_storno_sum_prije = 0
    total_iznos_sum_prije = 0
    uplata_sum_prije = 0
    user_data_list=[]
    for user_id in user_ids:
        contracts_sum_valid = Ugovor.objects.filter(
            storno=False,
            ugovor_zakljucen=True,
            korisnik=user_id,
            datum__gte=startdate,
            datum__lte=enddate
        ).aggregate(total_iznos=Sum('cijena'))

        contracts_sum_valid_prije = Ugovor.objects.filter(
            storno=False,
            ugovor_zakljucen=True,
            korisnik=user_id,
            datum__lt=startdate,
        ).aggregate(total_iznos_prije=Sum('cijena'))

        contracts_sum_storno = Ugovor.objects.filter(
            storno=True,
            ugovor_zakljucen=True,
            korisnik=user_id,
            datum__gte=startdate,
            datum__lte=enddate
        ).aggregate(total_iznos_storno=Sum('cijena'))

        contracts_sum_storno_prije = Ugovor.objects.filter(
            storno=True,
            ugovor_zakljucen=True,
            korisnik=user_id,
            datum__lt=startdate,
        ).aggregate(total_iznos_storno_prije=Sum('cijena'))

        uplata_sum = StavkeBankovnogIzvoda.objects.filter(
            id_bankovni_izvod__datum_izvoda__gte=startdate,
            id_bankovni_izvod__datum_izvoda__lte=enddate,
            korisnik_uplatio=user_id
        ).aggregate(total_uplata=Sum('uplata'))['total_uplata'] or 0

        uplata_sum_prije += StavkeBankovnogIzvoda.objects.filter(
            id_bankovni_izvod__datum_izvoda__lt=startdate,
            korisnik_uplatio=user_id
        ).aggregate(total_uplata=Sum('uplata'))['total_uplata'] or 0

        isplata_sum = StavkeBankovnogIzvoda.objects.filter(
            id_bankovni_izvod__datum_izvoda__gte=startdate,
            id_bankovni_izvod__datum_izvoda__lte=enddate,
            korisnik_uplatio=user_id
        ).aggregate(total_isplata=Sum('isplata'))['total_isplata'] or 0
        
        isplata_sum_prije = StavkeBankovnogIzvoda.objects.filter(
            
            id_bankovni_izvod__datum_izvoda__lt=startdate,
            korisnik_uplatio=user_id
        ).aggregate(total_isplata=Sum('isplata'))['total_isplata'] or 0

        total_iznos_sum = contracts_sum_valid.get('total_iznos', 0)
        total_storno_sum = contracts_sum_storno.get('total_iznos_storno', 0)
        total_iznos_sum_prije = contracts_sum_valid_prije.get('total_iznos_prije', 0)
        total_storno_sum_prije = contracts_sum_storno_prije.get('total_iznos_storno_prije', 0)

        if total_iznos_sum is None:
            total_iznos_sum=0
        if total_storno_sum is None:
            total_storno_sum=0
        if total_iznos_sum_prije is None:
            total_iznos_sum_prije = 0
        if total_storno_sum_prije is None:
            total_storno_sum_prije = 0
    
        uplatiti=total_iznos_sum + total_iznos_sum_prije -float(uplata_sum) - float(uplata_sum_prije)
        
        
        user_data= {
        "user_id": user_id,
        
        "total_iznos_sum": total_iznos_sum,
        "total_storno_sum": total_storno_sum,
        "uplata_sum": uplata_sum,
        "isplata_sum": isplata_sum,
        "isplata_sum_prije": isplata_sum_prije,
        "uplatiti": uplatiti,
        "start": startdate,
        "end": enddate,
        "uplata_sum_prije": uplata_sum_prije,
        "total_iznos_sum_prije": total_iznos_sum_prije,
        "contracts_sum_storno_prije": total_storno_sum_prije
        }
        user_data_list.append(user_data)
       
    startdate = parse_date(startdate)
    enddate = parse_date(enddate)
    
    context = {
        "user_data_list": user_data_list,
        "start": startdate,
        "end": enddate,
        "firma": preduzece,
        "base_dir": BASE_DIR,
    }
    pdf = render_to_pdf('kupoprodaja/rekapitulacija_korisnici.html', context)

    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = "rekapitulacija_%s.pdf" % ("12341231")
        content = "inline; filename='%s'" % (filename)
        download = request.GET.get("download")
        if download:
            content = "attachment; filename='%s'" % (filename)
        response['Content-Disposition'] = content
        return response
    return HttpResponse("Not found")
